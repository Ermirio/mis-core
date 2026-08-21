from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook
from tablib import Dataset

from equipamentos.models import (
    Area,
    ConexaoOPC,
    Equipamento,
    Fabrica,
    LinhaProducao,
    Sensor,
    TagColeta,
)
from equipamentos.resources import (
    AreaResource,
    EquipamentoConfigResource,
    EquipamentoResource,
    EquipamentoVariaveisWorkbook,
    LinhaProducaoResource,
    SensorResource,
    TagColetaResource,
)


class EquipamentoResourceImportTests(TestCase):
    def setUp(self):
        self.linha_01 = LinhaProducao.objects.create(
            codigo='L01',
            nome='Linha 01',
            localizacao='Galpao A',
            velocidade_planejada=100,
            meta_producao_hora=6000,
            meta_producao_turno=48000,
        )
        self.linha_02 = LinhaProducao.objects.create(
            codigo='L02',
            nome='Linha 02',
            localizacao='Galpao B',
            velocidade_planejada=100,
            meta_producao_hora=6000,
            meta_producao_turno=48000,
        )

    def _dataset(self, rows):
        dataset = Dataset()
        dataset.headers = [
            'codigo', 'nome', 'linha', 'tipo', 'ordem_na_linha', 'localizacao',
            'status', 'velocidade_nominal', 'velocidade_maxima', 'meta_oee',
            'temperatura_min', 'temperatura_max', 'pressao_min', 'pressao_max',
        ]
        for row in rows:
            dataset.append(row)
        return dataset

    def test_importa_mesmo_codigo_em_linhas_diferentes_sem_sobrescrever(self):
        dataset = self._dataset([
            ('E001', 'VINCULADORA - L01', 'L01', 'VINCULADORA', 1, '', 'ATIVO', 100, 120, 85, '', '', '', ''),
            ('E001', 'VINCULADORA - L02', 'L02', 'VINCULADORA', 1, '', 'ATIVO', 100, 120, 85, '', '', '', ''),
        ])

        result = EquipamentoResource().import_data(dataset, dry_run=False, raise_errors=True)

        self.assertFalse(result.has_errors())
        self.assertEqual(Equipamento.objects.count(), 2)
        self.assertTrue(
            Equipamento.objects.filter(
                linha=self.linha_01,
                codigo='E001',
                nome='VINCULADORA - L01',
            ).exists()
        )
        self.assertTrue(
            Equipamento.objects.filter(
                linha=self.linha_02,
                codigo='E001',
                nome='VINCULADORA - L02',
            ).exists()
        )

    def test_importa_linha_no_formato_antigo_com_codigo_e_nome(self):
        dataset = self._dataset([
            ('E002', 'ENCHEDORA - L01', 'L01 - Linha 01', 'ENCHEDORA', 2, '', 'ATIVO', 100, 120, 85, '', '', '', ''),
        ])

        result = EquipamentoResource().import_data(dataset, dry_run=False, raise_errors=True)

        self.assertFalse(result.has_errors())
        self.assertTrue(
            Equipamento.objects.filter(
                linha=self.linha_01,
                codigo='E002',
                nome='ENCHEDORA - L01',
            ).exists()
        )

    def test_importa_mesma_linha_e_codigo_como_atualizacao(self):
        Equipamento.objects.create(
            linha=self.linha_01,
            codigo='E003',
            nome='Nome antigo',
            tipo='ENCHEDORA',
            ordem_na_linha=3,
            velocidade_nominal=100,
            velocidade_maxima=120,
        )
        dataset = self._dataset([
            ('E003', 'Nome atualizado', 'L01', 'ENCHEDORA', 3, '', 'ATIVO', 100, 120, 85, '', '', '', ''),
        ])

        result = EquipamentoResource().import_data(dataset, dry_run=False, raise_errors=True)

        self.assertFalse(result.has_errors())
        self.assertEqual(Equipamento.objects.count(), 1)
        equipamento = Equipamento.objects.get(linha=self.linha_01, codigo='E003')
        self.assertEqual(equipamento.nome, 'Nome atualizado')

    def test_exporta_linha_por_codigo_e_nao_por_id_interno(self):
        Equipamento.objects.create(
            linha=self.linha_01,
            codigo='E004',
            nome='Exportado',
            tipo='ENCHEDORA',
            ordem_na_linha=4,
            velocidade_nominal=100,
            velocidade_maxima=120,
        )

        dataset = EquipamentoResource().export()
        row = dataset.dict[0]

        self.assertEqual(row['linha'], 'L01')


class TagColetaResourceImportTests(TestCase):
    def setUp(self):
        self.linha_01 = LinhaProducao.objects.create(
            codigo='L01',
            nome='Linha 01',
            localizacao='Galpao A',
            velocidade_planejada=100,
            meta_producao_hora=6000,
            meta_producao_turno=48000,
        )
        self.linha_02 = LinhaProducao.objects.create(
            codigo='L02',
            nome='Linha 02',
            localizacao='Galpao B',
            velocidade_planejada=100,
            meta_producao_hora=6000,
            meta_producao_turno=48000,
        )
        self.eq_l01 = Equipamento.objects.create(
            linha=self.linha_01,
            codigo='E001',
            nome='Enchedora L01',
            tipo='ENCHEDORA',
            velocidade_nominal=100,
            velocidade_maxima=120,
        )
        self.eq_l02 = Equipamento.objects.create(
            linha=self.linha_02,
            codigo='E001',
            nome='Enchedora L02',
            tipo='ENCHEDORA',
            velocidade_nominal=100,
            velocidade_maxima=120,
        )

    def _dataset(self, rows):
        dataset = Dataset()
        dataset.headers = [
            'linha', 'equipamento_codigo', 'equipamento', 'nome_metrica',
            'node_id', 'tipo_dado', 'formato', 'unidade', 'fator_conversao',
            'ativa', 'golden_state',
        ]
        for row in rows:
            dataset.append(row)
        return dataset

    def test_importa_mesma_variavel_em_equipamentos_iguais_de_linhas_diferentes(self):
        dataset = self._dataset([
            ('L01', 'E001', '', 'densidade', 'ns=2;s=L01.E001.Densidade', 'FLOAT', '', 'kg/m3', 1, True, False),
            ('L02', 'E001', '', 'densidade', 'ns=2;s=L02.E001.Densidade', 'FLOAT', '', 'kg/m3', 1, True, False),
        ])

        result = TagColetaResource().import_data(dataset, dry_run=False, raise_errors=True)

        self.assertFalse(result.has_errors())
        self.assertEqual(
            TagColeta.objects.get(equipamento=self.eq_l01, nome_metrica='densidade').node_id,
            'ns=2;s=L01.E001.Densidade',
        )
        self.assertEqual(
            TagColeta.objects.get(equipamento=self.eq_l02, nome_metrica='densidade').node_id,
            'ns=2;s=L02.E001.Densidade',
        )

    def test_importa_multiplas_variaveis_com_node_id_vazio_sem_colapsar(self):
        dataset = self._dataset([
            ('L01', 'E001', '', 'densidade', '', 'FLOAT', '', 'kg/m3', 1, False, False),
            ('L01', 'E001', '', 'temperatura', '', 'FLOAT', '', 'C', 1, False, False),
        ])

        result = TagColetaResource().import_data(dataset, dry_run=False, raise_errors=True)

        self.assertFalse(result.has_errors())
        self.assertTrue(TagColeta.objects.filter(equipamento=self.eq_l01, nome_metrica='densidade').exists())
        self.assertTrue(TagColeta.objects.filter(equipamento=self.eq_l01, nome_metrica='temperatura').exists())

    def test_exporta_tag_com_linha_codigo_e_slug_do_equipamento(self):
        tag = TagColeta.objects.create(
            equipamento=self.eq_l01,
            nome_metrica='pressao',
            node_id='ns=2;s=L01.E001.Pressao',
            tipo_dado='FLOAT',
            unidade='bar',
        )

        dataset = TagColetaResource().export(TagColeta.objects.filter(pk=tag.pk))
        row = dataset.dict[0]

        self.assertEqual(row['linha'], 'L01')
        self.assertEqual(row['equipamento_codigo'], 'E001')
        self.assertEqual(row['equipamento'], 'L01.E001')


class EquipamentoConfigResourceTests(TestCase):
    def setUp(self):
        self.linha_01 = LinhaProducao.objects.create(
            codigo='L01',
            nome='Linha 01',
            localizacao='Galpao A',
            velocidade_planejada=100,
            meta_producao_hora=6000,
            meta_producao_turno=48000,
        )
        self.linha_02 = LinhaProducao.objects.create(
            codigo='L02',
            nome='Linha 02',
            localizacao='Galpao B',
            velocidade_planejada=100,
            meta_producao_hora=6000,
            meta_producao_turno=48000,
        )

    def _dataset(self, rows):
        dataset = Dataset(headers=EquipamentoConfigResource.headers)
        for row in rows:
            dataset.append(row)
        return dataset

    def test_importa_config_com_mesmo_codigo_e_variavel_em_linhas_diferentes(self):
        dataset = self._dataset([
            ('', '', '', '', 'L01', 'Linha 01', 'E001', 'Enchedora L01', '', 'ENCHEDORA', 1, '', 'ATIVO', 100, 120, 85, 'densidade', 'ns=2;s=L01.E001.Densidade', 'FLOAT', '', 'kg/m3', 1, True, False),
            ('', '', '', '', 'L02', 'Linha 02', 'E001', 'Enchedora L02', '', 'ENCHEDORA', 1, '', 'ATIVO', 100, 120, 85, 'densidade', 'ns=2;s=L02.E001.Densidade', 'FLOAT', '', 'kg/m3', 1, True, False),
        ])

        summary = EquipamentoConfigResource().import_config_data(dataset, dry_run=False)

        self.assertEqual(summary['errors'], [])
        self.assertEqual(Equipamento.objects.filter(codigo='E001').count(), 2)
        self.assertEqual(
            TagColeta.objects.get(
                equipamento__linha=self.linha_01,
                equipamento__codigo='E001',
                nome_metrica='densidade',
            ).node_id,
            'ns=2;s=L01.E001.Densidade',
        )
        self.assertEqual(
            TagColeta.objects.get(
                equipamento__linha=self.linha_02,
                equipamento__codigo='E001',
                nome_metrica='densidade',
            ).node_id,
            'ns=2;s=L02.E001.Densidade',
        )

    def test_validacao_dry_run_nao_salva(self):
        dataset = self._dataset([
            ('', '', '', '', 'L01', 'Linha 01', 'E001', 'Enchedora L01', '', 'ENCHEDORA', 1, '', 'ATIVO', 100, 120, 85, 'densidade', 'ns=2;s=L01.E001.Densidade', 'FLOAT', '', 'kg/m3', 1, True, False),
        ])

        summary = EquipamentoConfigResource().import_config_data(dataset, dry_run=True)

        self.assertEqual(summary['errors'], [])
        self.assertEqual(Equipamento.objects.count(), 0)
        self.assertEqual(TagColeta.objects.count(), 0)

    def test_exporta_config_com_hierarquia_e_variavel(self):
        equipamento = Equipamento.objects.create(
            linha=self.linha_01,
            codigo='E001',
            nome='Enchedora L01',
            tipo='ENCHEDORA',
            velocidade_nominal=100,
            velocidade_maxima=120,
        )
        tag = TagColeta.objects.create(
            equipamento=equipamento,
            nome_metrica='densidade',
            node_id='ns=2;s=L01.E001.Densidade',
            tipo_dado='FLOAT',
            unidade='kg/m3',
        )

        dataset = EquipamentoConfigResource().export(
            queryset=Equipamento.objects.filter(pk=equipamento.pk)
        )
        rows = [row for row in dataset.dict if row['nome_metrica'] == tag.nome_metrica]

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['linha_codigo'], 'L01')
        self.assertEqual(rows[0]['equipamento_codigo'], 'E001')
        self.assertEqual(rows[0]['equipamento_slug'], 'L01.E001')
        self.assertEqual(rows[0]['node_id'], 'ns=2;s=L01.E001.Densidade')


@override_settings(STORAGES={
    'default': {'BACKEND': 'django.core.files.storage.FileSystemStorage'},
    'staticfiles': {'BACKEND': 'django.contrib.staticfiles.storage.StaticFilesStorage'},
})
class EquipamentoConfigAdminViewsTests(TestCase):
    def setUp(self):
        user = get_user_model().objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='admin123',
        )
        self.client.force_login(user)
        self.linha = LinhaProducao.objects.create(
            codigo='L01',
            nome='Linha 01',
            localizacao='Galpao A',
            velocidade_planejada=100,
            meta_producao_hora=6000,
            meta_producao_turno=48000,
        )
        self.equipamento = Equipamento.objects.create(
            linha=self.linha,
            codigo='E001',
            nome='Enchedora',
            tipo='ENCHEDORA',
            velocidade_nominal=100,
            velocidade_maxima=120,
        )

    def test_changelist_exibe_um_botao_adicionar_na_barra_de_acoes(self):
        response = self.client.get(reverse('admin:equipamentos_equipamento_changelist'))

        self.assertEqual(response.status_code, 200)
        add_url = reverse('admin:equipamentos_equipamento_add')
        html = response.content.decode()
        object_tools = html.split('<ul class="object-tools">', 1)[1].split('</ul>', 1)[0]
        self.assertEqual(
            object_tools.count(f'href="{add_url}" class="addlink"'),
            1,
        )

    def test_change_form_exibe_acoes_de_variaveis(self):
        response = self.client.get(reverse(
            'admin:equipamentos_equipamento_change',
            args=[self.equipamento.pk],
        ))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Exportar variáveis', count=1)
        self.assertContains(response, 'Importar variáveis', count=1)

    def test_export_individual_retorna_tres_abas_sem_outro_equipamento(self):
        outra_linha = LinhaProducao.objects.create(
            codigo='L02',
            nome='Linha 02',
            localizacao='Galpao B',
            velocidade_planejada=100,
            meta_producao_hora=6000,
            meta_producao_turno=48000,
        )
        outro = Equipamento.objects.create(
            linha=outra_linha,
            codigo='E001',
            nome='Enchedora L02',
            tipo='ENCHEDORA',
            velocidade_nominal=100,
            velocidade_maxima=120,
        )
        TagColeta.objects.create(
            equipamento=self.equipamento,
            nome_metrica='densidade',
            node_id='ns=2;s=L01.E001.Densidade',
            tipo_dado='FLOAT',
        )
        TagColeta.objects.create(
            equipamento=outro,
            nome_metrica='densidade',
            node_id='ns=2;s=L02.E001.Densidade',
            tipo_dado='FLOAT',
        )

        response = self.client.get(reverse(
            'admin:equipamentos_equipamento_variables_export',
            args=[self.equipamento.pk],
        ))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        self.assertEqual(workbook.sheetnames, [
            EquipamentoVariaveisWorkbook.SHEET_DEFAULT,
            EquipamentoVariaveisWorkbook.SHEET_SENSORS,
            EquipamentoVariaveisWorkbook.SHEET_ANALYTICS,
        ])
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                self.assertEqual(row[0], 'L01')
                self.assertEqual(row[1], 'E001')
                self.assertNotIn('L02', row)

        analytics = workbook[EquipamentoVariaveisWorkbook.SHEET_ANALYTICS]
        self.assertEqual(analytics['C2'].value, 'densidade')
        self.assertEqual(analytics['D2'].value, 'ns=2;s=L01.E001.Densidade')

    def test_import_individual_cria_variavel_e_sensor(self):
        content = EquipamentoVariaveisWorkbook().export(self.equipamento)
        workbook = load_workbook(BytesIO(content))

        analytics = workbook[EquipamentoVariaveisWorkbook.SHEET_ANALYTICS]
        analytics.append((
            'L01', 'E001', 'densidade', 'ns=2;s=L01.E001.Densidade',
            'FLOAT', '', 'kg/m3', 1, True, False,
        ))
        sensors = workbook[EquipamentoVariaveisWorkbook.SHEET_SENSORS]
        sensors.append((
            'L01', 'E001', '', 'Temperatura tanque', 'INPUT_FLOAT',
            'temperatura_tanque', 'C', True, False, 0, 100, 10, 90, 50, '',
        ))
        output = BytesIO()
        workbook.save(output)
        upload = SimpleUploadedFile(
            'variaveis-L01-E001.xlsx',
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response = self.client.post(
            reverse(
                'admin:equipamentos_equipamento_variables_import',
                args=[self.equipamento.pk],
            ),
            {'arquivo': upload, 'importar': '1'},
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(TagColeta.objects.filter(
            equipamento=self.equipamento,
            nome_metrica='densidade',
            node_id='ns=2;s=L01.E001.Densidade',
        ).exists())
        self.assertTrue(Sensor.objects.filter(
            equipamento=self.equipamento,
            nome='Temperatura tanque',
            codigo='S001',
        ).exists())

    def test_import_de_outro_equipamento_faz_rollback_total(self):
        outra_linha = LinhaProducao.objects.create(
            codigo='L02',
            nome='Linha 02',
            localizacao='Galpao B',
            velocidade_planejada=100,
            meta_producao_hora=6000,
            meta_producao_turno=48000,
        )
        outro = Equipamento.objects.create(
            linha=outra_linha,
            codigo='E001',
            nome='Enchedora L02',
            tipo='ENCHEDORA',
            velocidade_nominal=100,
            velocidade_maxima=120,
        )
        content = EquipamentoVariaveisWorkbook().export(self.equipamento)
        upload = SimpleUploadedFile(
            'variaveis-L01-E001.xlsx',
            content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response = self.client.post(
            reverse(
                'admin:equipamentos_equipamento_variables_import',
                args=[outro.pk],
            ),
            {'arquivo': upload, 'importar': '1'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'nao corresponde')
        self.assertFalse(
            outro.tags_coleta.exclude(node_id='').exists()
        )

    def test_import_nao_exclui_variavel_ausente_da_planilha(self):
        existente = TagColeta.objects.create(
            equipamento=self.equipamento,
            nome_metrica='pressao_custom',
            node_id='ns=2;s=Pressao',
            tipo_dado='FLOAT',
        )
        content = EquipamentoVariaveisWorkbook().export(self.equipamento)
        workbook = load_workbook(BytesIO(content))
        analytics = workbook[EquipamentoVariaveisWorkbook.SHEET_ANALYTICS]
        analytics.delete_rows(2, analytics.max_row - 1)
        output = BytesIO()
        workbook.save(output)
        upload = SimpleUploadedFile(
            'variaveis-L01-E001.xlsx',
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response = self.client.post(
            reverse(
                'admin:equipamentos_equipamento_variables_import',
                args=[self.equipamento.pk],
            ),
            {'arquivo': upload, 'importar': '1'},
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(TagColeta.objects.filter(pk=existente.pk).exists())


class HierarquiaResourceIdentityTests(TestCase):
    def test_area_exporta_e_importa_fabrica_por_codigo(self):
        fabrica = Fabrica.objects.create(codigo='F001', nome='Fabrica 01')
        area = Area.objects.create(codigo='A01', nome='Envase', fabrica=fabrica)

        exported = AreaResource().export(Area.objects.filter(pk=area.pk)).dict[0]
        self.assertEqual(exported['fabrica'], 'F001')

        Area.objects.all().delete()
        dataset = Dataset(headers=['codigo', 'nome', 'fabrica'])
        dataset.append(('A01', 'Envase Novo', 'F001'))
        result = AreaResource().import_data(dataset, dry_run=False, raise_errors=True)

        self.assertFalse(result.has_errors())
        self.assertEqual(Area.objects.get(codigo='A01').fabrica.codigo, 'F001')

    def test_linha_exporta_area_e_conexao_por_codigo_ou_nome_estavel(self):
        fabrica = Fabrica.objects.create(codigo='F001', nome='Fabrica 01')
        area = Area.objects.create(codigo='A01', nome='Envase', fabrica=fabrica)
        conexao = ConexaoOPC.objects.create(
            nome='OPC-L01',
            url_servidor='opc.tcp://localhost:4840',
        )
        linha = LinhaProducao.objects.create(
            codigo='L01',
            nome='Linha 01',
            area=area,
            conexao_padrao=conexao,
            localizacao='Galpao A',
            velocidade_planejada=100,
            meta_producao_hora=6000,
            meta_producao_turno=48000,
        )

        exported = LinhaProducaoResource().export(
            LinhaProducao.objects.filter(pk=linha.pk)
        ).dict[0]

        self.assertEqual(exported['area'], 'A01')
        self.assertEqual(exported['conexao_padrao'], 'OPC-L01')

    def test_sensor_importa_mesmo_codigo_em_equipamentos_diferentes(self):
        linha = LinhaProducao.objects.create(
            codigo='L01',
            nome='Linha 01',
            localizacao='Galpao A',
            velocidade_planejada=100,
            meta_producao_hora=6000,
            meta_producao_turno=48000,
        )
        eq1 = Equipamento.objects.create(
            linha=linha,
            codigo='E001',
            nome='Enchedora',
            tipo='ENCHEDORA',
            velocidade_nominal=100,
            velocidade_maxima=120,
        )
        eq2 = Equipamento.objects.create(
            linha=linha,
            codigo='E002',
            nome='Paletizadora',
            tipo='PALETIZADORA',
            velocidade_nominal=100,
            velocidade_maxima=120,
        )
        dataset = Dataset(headers=[
            'codigo', 'nome', 'tipo', 'tag_influxdb', 'unidade',
            'linha', 'equipamento', 'valor_min', 'valor_max', 'ativo',
        ])
        dataset.append(('S001', 'Sensor A', 'INPUT_FLOAT', 'pressao_a', 'bar', '', eq1.slug, '', '', True))
        dataset.append(('S001', 'Sensor B', 'INPUT_FLOAT', 'pressao_b', 'bar', '', eq2.slug, '', '', True))

        result = SensorResource().import_data(dataset, dry_run=False, raise_errors=True)

        self.assertFalse(result.has_errors())
        self.assertTrue(Sensor.objects.filter(equipamento=eq1, codigo='S001').exists())
        self.assertTrue(Sensor.objects.filter(equipamento=eq2, codigo='S001').exists())
