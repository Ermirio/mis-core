from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.core.exceptions import ValidationError
from django.db import transaction
from import_export import fields, resources
from import_export.widgets import ForeignKeyWidget
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from tablib import Dataset

from .models import (
    Equipamento, Fabrica, Area, Produto, LinhaProducao,
    TagColeta, Sensor, ConexaoOPC, OrdemProducao, TurnoProducao,
    DEFAULT_TAGS_COLETA, DEFAULT_TAGS_BY_NAME,
)


class CodigoOuPkLegadoWidget(ForeignKeyWidget):
    """Resolve FK por campo estavel e ainda aceita PK numerica legada."""

    def __init__(self, model, field='codigo'):
        self.lookup_field = field
        super().__init__(model, field)

    def clean(self, value, row=None, **kwargs):
        raw = '' if value is None else str(value).strip()
        if not raw:
            return None

        lookup_value = raw.split(' - ', 1)[0].strip()
        obj = self.model.objects.filter(**{self.lookup_field: lookup_value}).first()
        if obj:
            return obj

        if raw.isdigit():
            obj = self.model.objects.filter(pk=int(raw)).first()
            if obj:
                return obj

        return super().clean(raw, row=row, **kwargs)

    def render(self, value, obj=None, **kwargs):
        return getattr(value, self.lookup_field) if value else ''


class LinhaProducaoWidget(CodigoOuPkLegadoWidget):
    """Resolve linha por codigo, mantendo compatibilidade com exports antigos.

    O import/export de Equipamento precisa usar a identidade real do cadastro:
    linha + codigo do equipamento. Sem isso, E001 de L01 e E001 de L02 viram o
    mesmo registro durante o import e a ultima linha da planilha sobrescreve as
    anteriores.
    """

    def __init__(self):
        super().__init__(LinhaProducao, 'codigo')


class FabricaWidget(CodigoOuPkLegadoWidget):
    def __init__(self):
        super().__init__(Fabrica, 'codigo')


class AreaWidget(CodigoOuPkLegadoWidget):
    def __init__(self):
        super().__init__(Area, 'codigo')


class ProdutoWidget(CodigoOuPkLegadoWidget):
    def __init__(self):
        super().__init__(Produto, 'codigo')


class ConexaoOPCWidget(CodigoOuPkLegadoWidget):
    def __init__(self):
        super().__init__(ConexaoOPC, 'nome')


class EquipamentoHierarquicoWidget(ForeignKeyWidget):
    """Resolve equipamento por hierarquia, nao por id interno do banco."""

    def __init__(self):
        super().__init__(Equipamento, 'slug')

    def _normalizar(self, value):
        raw = '' if value is None else str(value).strip()
        return raw.split(' - ', 1)[0].strip()

    def clean(self, value, row=None, **kwargs):
        row = row or {}
        linha_codigo = self._normalizar(row.get('linha'))
        equipamento_codigo = self._normalizar(
            row.get('equipamento_codigo')
            or row.get('codigo_equipamento')
            or row.get('codigo')
        )
        raw = self._normalizar(value)

        if linha_codigo and equipamento_codigo:
            return Equipamento.objects.select_related('linha').get(
                linha__codigo=linha_codigo,
                codigo=equipamento_codigo,
            )

        if raw:
            equipamento = Equipamento.objects.filter(slug=raw).first()
            if equipamento:
                return equipamento

            if linha_codigo:
                equipamento = Equipamento.objects.filter(
                    linha__codigo=linha_codigo,
                    codigo=raw,
                ).first()
                if equipamento:
                    return equipamento

            if '.' in raw:
                linha_from_slug, codigo_from_slug = raw.split('.', 1)
                equipamento = Equipamento.objects.filter(
                    linha__codigo=linha_from_slug,
                    codigo=codigo_from_slug,
                ).first()
                if equipamento:
                    return equipamento

            if raw.isdigit():
                equipamento = Equipamento.objects.filter(pk=int(raw)).first()
                if equipamento:
                    return equipamento

        return super().clean(raw, row=row, **kwargs)

    def render(self, value, obj=None, **kwargs):
        if not value:
            return ''
        return value.slug or f'{value.linha.codigo}.{value.codigo}'


class EquipamentoResource(resources.ModelResource):
    linha = fields.Field(
        column_name='linha',
        attribute='linha',
        widget=LinhaProducaoWidget(),
    )

    class Meta:
        model = Equipamento
        import_id_fields = ('linha', 'codigo')
        fields = (
            'codigo', 'nome', 'linha', 'tipo', 'ordem_na_linha', 'localizacao',
            'status', 'velocidade_nominal', 'velocidade_maxima', 'meta_oee',
            'temperatura_min', 'temperatura_max', 'pressao_min', 'pressao_max',
        )

class FabricaResource(resources.ModelResource):
    class Meta:
        model = Fabrica
        import_id_fields = ('codigo',)
        fields = ('codigo', 'nome', 'localizacao')

class AreaResource(resources.ModelResource):
    fabrica = fields.Field(
        column_name='fabrica',
        attribute='fabrica',
        widget=FabricaWidget(),
    )

    class Meta:
        model = Area
        import_id_fields = ('codigo',)
        fields = ('codigo', 'nome', 'fabrica')

class ProdutoResource(resources.ModelResource):
    class Meta:
        model = Produto
        import_id_fields = ('codigo',)
        fields = ('codigo', 'descricao', 'peso_unitario', 'ativo')

class LinhaProducaoResource(resources.ModelResource):
    area = fields.Field(
        column_name='area',
        attribute='area',
        widget=AreaWidget(),
    )
    conexao_padrao = fields.Field(
        column_name='conexao_padrao',
        attribute='conexao_padrao',
        widget=ConexaoOPCWidget(),
    )

    class Meta:
        model = LinhaProducao
        import_id_fields = ('codigo',)
        fields = (
            'codigo', 'nome', 'area', 'conexao_padrao', 'descricao', 'localizacao', 'ativa',
            'velocidade_planejada', 'meta_producao_hora', 'meta_producao_turno', 'meta_oee'
        )

class TagColetaResource(resources.ModelResource):
    linha = fields.Field(
        column_name='linha',
        readonly=True,
    )
    equipamento_codigo = fields.Field(
        column_name='equipamento_codigo',
        readonly=True,
    )
    equipamento = fields.Field(
        column_name='equipamento',
        attribute='equipamento',
        widget=EquipamentoHierarquicoWidget(),
    )

    def dehydrate_linha(self, obj):
        return obj.equipamento.linha.codigo if obj.equipamento_id else ''

    def dehydrate_equipamento_codigo(self, obj):
        return obj.equipamento.codigo if obj.equipamento_id else ''

    class Meta:
        model = TagColeta
        import_id_fields = ('equipamento', 'nome_metrica')
        fields = (
            'linha', 'equipamento_codigo', 'equipamento', 'nome_metrica',
            'node_id', 'tipo_dado', 'formato', 'unidade', 'fator_conversao',
            'ativa', 'golden_state',
        )


class EquipamentoConfigResource(resources.Resource):
    """Planilha operacional: uma linha por variavel de equipamento.

    Identidade usada no import:
    - linha_codigo + equipamento_codigo para o equipamento
    - equipamento + nome_metrica para a variavel

    Campos de nome (fabrica_nome, area_nome, linha_nome, equipamento_nome) sao
    contexto humano, nao chaves. Isso evita sobrescrita quando nomes mudam ou
    quando E001 existe em mais de uma linha.
    """

    headers = [
        'fabrica_codigo',
        'fabrica_nome',
        'area_codigo',
        'area_nome',
        'linha_codigo',
        'linha_nome',
        'equipamento_codigo',
        'equipamento_nome',
        'equipamento_slug',
        'tipo',
        'ordem_na_linha',
        'localizacao',
        'status',
        'velocidade_nominal',
        'velocidade_maxima',
        'meta_oee',
        'nome_metrica',
        'node_id',
        'tipo_dado',
        'formato',
        'unidade',
        'fator_conversao',
        'ativa',
        'golden_state',
    ]

    TRUE_VALUES = {'1', 'true', 't', 'sim', 's', 'yes', 'y', 'ativo', 'ativa'}
    FALSE_VALUES = {'0', 'false', 'f', 'nao', 'não', 'n', 'no', 'inativo', 'inativa'}

    def export(self, *args, queryset=None, **kwargs):
        dataset = Dataset(headers=self.headers)
        qs = queryset or Equipamento.objects.all()
        qs = qs.select_related('linha__area__fabrica').prefetch_related('tags_coleta')
        qs = qs.order_by('linha__codigo', 'ordem_na_linha', 'codigo')

        for equipamento in qs:
            tags = list(equipamento.tags_coleta.all().order_by('nome_metrica'))
            if tags:
                for tag in tags:
                    dataset.append(self._row_for(equipamento, tag))
            else:
                dataset.append(self._row_for(equipamento, None))
        return dataset

    def import_config_data(self, dataset, dry_run=False):
        summary = {
            'total_rows': len(dataset),
            'processed_rows': 0,
            'equipment_created': 0,
            'equipment_updated': 0,
            'tags_created': 0,
            'tags_updated': 0,
            'errors': [],
            'dry_run': dry_run,
        }
        created_equipment = set()
        updated_equipment = set()

        with transaction.atomic():
            for idx, row in enumerate(dataset.dict, start=2):
                if self._is_blank_row(row):
                    continue

                try:
                    with transaction.atomic():
                        eq, eq_created = self._upsert_equipment(row)
                        eq_key = (eq.linha_id, eq.codigo)
                        if eq_created:
                            created_equipment.add(eq_key)
                        else:
                            updated_equipment.add(eq_key)

                        tag_result = self._upsert_tag(row, eq)
                        if tag_result == 'created':
                            summary['tags_created'] += 1
                        elif tag_result == 'updated':
                            summary['tags_updated'] += 1

                        summary['processed_rows'] += 1
                except Exception as exc:
                    summary['errors'].append({
                        'row': idx,
                        'error': str(exc),
                    })

            if summary['errors'] or dry_run:
                transaction.set_rollback(True)

        summary['equipment_created'] = len(created_equipment)
        summary['equipment_updated'] = len(updated_equipment - created_equipment)
        return summary

    @classmethod
    def dataset_from_upload(cls, uploaded_file):
        raw = uploaded_file.read()
        name = uploaded_file.name.lower()
        dataset = Dataset()

        if name.endswith('.xlsx'):
            dataset.load(raw, format='xlsx')
        elif name.endswith('.xls'):
            dataset.load(raw, format='xls')
        elif name.endswith('.csv'):
            dataset.load(raw.decode('utf-8-sig'), format='csv')
        elif name.endswith('.tsv'):
            dataset.load(raw.decode('utf-8-sig'), format='tsv')
        elif name.endswith('.json'):
            dataset.load(raw.decode('utf-8-sig'), format='json')
        else:
            raise ValidationError('Formato nao suportado. Use .xlsx, .xls, .csv, .tsv ou .json.')

        return dataset

    def _row_for(self, equipamento, tag):
        linha = equipamento.linha
        area = linha.area if linha else None
        fabrica = area.fabrica if area else None
        return [
            fabrica.codigo if fabrica else '',
            fabrica.nome if fabrica else '',
            area.codigo if area else '',
            area.nome if area else '',
            linha.codigo if linha else '',
            linha.nome if linha else '',
            equipamento.codigo,
            equipamento.nome,
            equipamento.slug,
            equipamento.tipo,
            equipamento.ordem_na_linha,
            equipamento.localizacao,
            equipamento.status,
            equipamento.velocidade_nominal,
            equipamento.velocidade_maxima,
            equipamento.meta_oee,
            tag.nome_metrica if tag else '',
            tag.node_id if tag else '',
            tag.tipo_dado if tag else '',
            tag.formato if tag else '',
            tag.unidade if tag else '',
            tag.fator_conversao if tag else '',
            tag.ativa if tag else '',
            tag.golden_state if tag else '',
        ]

    def _upsert_equipment(self, row):
        linha_codigo = self._get(row, 'linha_codigo', 'linha', 'codigo_linha')
        equipamento_slug = self._get(row, 'equipamento_slug', 'equipamento', 'slug')
        equipamento_codigo = self._get(
            row,
            'equipamento_codigo',
            'codigo_equipamento',
            'codigo',
        )

        if not linha_codigo and equipamento_slug and '.' in equipamento_slug:
            linha_codigo, equipamento_codigo_from_slug = equipamento_slug.split('.', 1)
            equipamento_codigo = equipamento_codigo or equipamento_codigo_from_slug

        if not linha_codigo:
            raise ValidationError('linha_codigo e obrigatorio.')
        if not equipamento_codigo:
            raise ValidationError('equipamento_codigo e obrigatorio.')

        linha = LinhaProducao.objects.filter(codigo=linha_codigo).first()
        if not linha:
            raise ValidationError(f'Linha "{linha_codigo}" nao encontrada. Cadastre/importe a linha antes.')

        equipamento = Equipamento.objects.filter(
            linha=linha,
            codigo=equipamento_codigo,
        ).first()
        created = equipamento is None
        if created:
            equipamento = Equipamento(linha=linha, codigo=equipamento_codigo)

        nome = self._get(row, 'equipamento_nome', 'nome_equipamento', 'nome')
        tipo = self._get(row, 'tipo', 'equipamento_tipo')
        status = self._get(row, 'status')

        if nome:
            equipamento.nome = nome
        elif created:
            equipamento.nome = equipamento_codigo

        if tipo:
            equipamento.tipo = tipo
        elif created:
            equipamento.tipo = 'OUTRO'

        if status:
            equipamento.status = status
        elif created:
            equipamento.status = 'ATIVO'

        equipamento.ordem_na_linha = self._to_int(
            self._get(row, 'ordem_na_linha', 'ordem'),
            equipamento.ordem_na_linha if not created else 1,
        )
        equipamento.localizacao = self._get(row, 'localizacao') or equipamento.localizacao
        equipamento.velocidade_nominal = self._to_float(
            self._get(row, 'velocidade_nominal'),
            equipamento.velocidade_nominal if not created else 0.0,
        )
        equipamento.velocidade_maxima = self._to_float(
            self._get(row, 'velocidade_maxima'),
            equipamento.velocidade_maxima if not created else 0.0,
        )
        equipamento.meta_oee = self._to_float(
            self._get(row, 'meta_oee'),
            equipamento.meta_oee if not created else 85.0,
        )
        equipamento.save()
        return equipamento, created

    def _upsert_tag(self, row, equipamento):
        nome_metrica = self._get(row, 'nome_metrica', 'variavel', 'variavel_nome', 'tag')
        if not nome_metrica:
            return 'skipped'

        tag = TagColeta.objects.filter(
            equipamento=equipamento,
            nome_metrica=nome_metrica,
        ).first()
        created = tag is None
        if created:
            tag = TagColeta(equipamento=equipamento, nome_metrica=nome_metrica)

        if self._has(row, 'node_id'):
            tag.node_id = self._get(row, 'node_id')
        if self._has(row, 'tipo_dado') and self._get(row, 'tipo_dado'):
            tag.tipo_dado = self._get(row, 'tipo_dado')
        if self._has(row, 'formato'):
            tag.formato = self._to_decimal(self._get(row, 'formato'), None)
        if self._has(row, 'unidade'):
            tag.unidade = self._get(row, 'unidade')
        if self._has(row, 'fator_conversao'):
            tag.fator_conversao = self._to_float(
                self._get(row, 'fator_conversao'),
                tag.fator_conversao,
            )
        if self._has(row, 'ativa'):
            tag.ativa = self._to_bool(self._get(row, 'ativa'), tag.ativa)
        if self._has(row, 'golden_state'):
            tag.golden_state = self._to_bool(self._get(row, 'golden_state'), tag.golden_state)

        tag.save()
        return 'created' if created else 'updated'

    def _get(self, row, *names):
        for name in names:
            if name in row:
                value = row.get(name)
                if value is None:
                    return ''
                return str(value).strip()
        return ''

    def _has(self, row, name):
        return name in row

    def _is_blank_row(self, row):
        return not any(str(value).strip() for value in row.values() if value is not None)

    def _to_float(self, value, default):
        if value in (None, ''):
            return default
        return float(str(value).replace(',', '.'))

    def _to_int(self, value, default):
        if value in (None, ''):
            return default
        return int(float(str(value).replace(',', '.')))

    def _to_decimal(self, value, default):
        if value in (None, ''):
            return default
        try:
            return Decimal(str(value).replace(',', '.'))
        except InvalidOperation as exc:
            raise ValidationError(f'Valor decimal invalido: {value}') from exc

    def _to_bool(self, value, default):
        if value in (None, ''):
            return default
        normalized = str(value).strip().lower()
        if normalized in self.TRUE_VALUES:
            return True
        if normalized in self.FALSE_VALUES:
            return False
        raise ValidationError(f'Valor booleano invalido: {value}')


class EquipamentoVariaveisWorkbook:
    """Importa e exporta a configuracao de um unico equipamento.

    O arquivo tem tres abas porque TagColeta e Sensor possuem contratos
    diferentes. A identidade da URL e conferida com linha_codigo e
    equipamento_codigo em cada linha para impedir importacao no equipamento
    errado. Registros ausentes nunca sao excluidos.
    """

    SHEET_DEFAULT = 'Variaveis padrao'
    SHEET_SENSORS = 'Sensores'
    SHEET_ANALYTICS = 'Analise e historico'
    SHEETS = (SHEET_DEFAULT, SHEET_SENSORS, SHEET_ANALYTICS)

    TAG_HEADERS = [
        'linha_codigo', 'equipamento_codigo', 'nome_metrica', 'node_id',
        'tipo_dado', 'formato', 'unidade', 'fator_conversao', 'ativa',
        'golden_state',
    ]
    SENSOR_HEADERS = [
        'linha_codigo', 'equipamento_codigo', 'codigo', 'nome', 'tipo',
        'tag_influxdb', 'unidade', 'ativo', 'golden_state', 'valor_min',
        'valor_max', 'lsl', 'usl', 'nominal', 'observacoes',
    ]
    TRUE_VALUES = {'1', 'true', 't', 'sim', 's', 'yes', 'y', 'ativo', 'ativa'}
    FALSE_VALUES = {'0', 'false', 'f', 'nao', 'não', 'n', 'no', 'inativo', 'inativa'}

    def export(self, equipamento):
        workbook = Workbook()
        workbook.remove(workbook.active)

        tags = {
            tag.nome_metrica: tag
            for tag in equipamento.tags_coleta.all().order_by('nome_metrica')
        }
        default_rows = []
        for defaults in DEFAULT_TAGS_COLETA:
            default_rows.append(
                self._tag_row(equipamento, tags.get(defaults['nome']), defaults)
            )

        custom_rows = [
            self._tag_row(equipamento, tag)
            for name, tag in tags.items()
            if name not in DEFAULT_TAGS_BY_NAME
        ]
        if not custom_rows:
            custom_rows.append(self._tag_row(equipamento, None))

        sensor_rows = [
            self._sensor_row(equipamento, sensor)
            for sensor in equipamento.sensores.all().order_by('codigo', 'nome')
        ]
        if not sensor_rows:
            sensor_rows.append(self._sensor_row(equipamento, None))

        self._add_sheet(
            workbook, self.SHEET_DEFAULT, self.TAG_HEADERS, default_rows,
            tab_color='2F75B5',
        )
        self._add_sheet(
            workbook, self.SHEET_SENSORS, self.SENSOR_HEADERS, sensor_rows,
            tab_color='70AD47',
        )
        self._add_sheet(
            workbook, self.SHEET_ANALYTICS, self.TAG_HEADERS, custom_rows,
            tab_color='C55A11',
        )

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def import_data(self, uploaded_file, equipamento, dry_run=False):
        name = uploaded_file.name.lower()
        if not name.endswith('.xlsx'):
            raise ValidationError('Formato nao suportado. Use o arquivo .xlsx exportado pelo sistema.')

        try:
            workbook = load_workbook(uploaded_file, data_only=True)
        except Exception as exc:
            raise ValidationError(f'Nao foi possivel abrir a planilha: {exc}') from exc

        present_sheets = [name for name in self.SHEETS if name in workbook.sheetnames]
        if not present_sheets:
            expected = ', '.join(self.SHEETS)
            raise ValidationError(f'Nenhuma aba reconhecida. Abas esperadas: {expected}.')

        summary = {
            'total_rows': 0,
            'processed_rows': 0,
            'tags_created': 0,
            'tags_updated': 0,
            'sensors_created': 0,
            'sensors_updated': 0,
            'errors': [],
            'dry_run': dry_run,
        }

        with transaction.atomic():
            for sheet_name in present_sheets:
                worksheet = workbook[sheet_name]
                expected_headers = (
                    self.SENSOR_HEADERS
                    if sheet_name == self.SHEET_SENSORS
                    else self.TAG_HEADERS
                )
                try:
                    rows = self._worksheet_rows(worksheet, expected_headers)
                except ValidationError as exc:
                    summary['errors'].append({
                        'sheet': sheet_name,
                        'row': 1,
                        'error': str(exc),
                    })
                    continue

                for row_number, row in rows:
                    payload_headers = expected_headers[2:]
                    if not any(self._text(row.get(key)) for key in payload_headers):
                        continue

                    summary['total_rows'] += 1
                    try:
                        with transaction.atomic():
                            self._validate_identity(row, equipamento)
                            if sheet_name == self.SHEET_SENSORS:
                                result = self._upsert_sensor(row, equipamento)
                                summary[f'sensors_{result}'] += 1
                            else:
                                result = self._upsert_tag(
                                    row,
                                    equipamento,
                                    standard=(sheet_name == self.SHEET_DEFAULT),
                                )
                                summary[f'tags_{result}'] += 1
                            summary['processed_rows'] += 1
                    except Exception as exc:
                        summary['errors'].append({
                            'sheet': sheet_name,
                            'row': row_number,
                            'error': str(exc),
                        })

            if summary['errors'] or dry_run:
                transaction.set_rollback(True)

        return summary

    def _add_sheet(self, workbook, title, headers, rows, tab_color):
        worksheet = workbook.create_sheet(title=title)
        worksheet.sheet_properties.tabColor = tab_color
        worksheet.freeze_panes = 'A2'
        worksheet.append(headers)
        for row in rows:
            worksheet.append(row)

        header_fill = PatternFill('solid', fgColor='1F4E78')
        for cell in worksheet[1]:
            cell.font = Font(color='FFFFFF', bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        worksheet.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{worksheet.max_row}'
        for column_index, header in enumerate(headers, start=1):
            values = [header]
            values.extend(
                '' if cell.value is None else str(cell.value)
                for cell in list(worksheet.columns)[column_index - 1][1:]
            )
            width = min(max(len(value) for value in values) + 2, 48)
            worksheet.column_dimensions[get_column_letter(column_index)].width = max(width, 12)

    def _tag_row(self, equipamento, tag, defaults=None):
        has_defaults = defaults is not None
        defaults = defaults or {}
        return [
            equipamento.linha.codigo,
            equipamento.codigo,
            tag.nome_metrica if tag else defaults.get('nome', ''),
            tag.node_id if tag else '',
            tag.tipo_dado if tag else defaults.get('tipo_dado', ''),
            tag.formato if tag else '',
            tag.unidade if tag else defaults.get('unidade', ''),
            tag.fator_conversao if tag else defaults.get('fator_conversao', '') if has_defaults else '',
            tag.ativa if tag else False if has_defaults else '',
            tag.golden_state if tag else False if has_defaults else '',
        ]

    def _sensor_row(self, equipamento, sensor):
        return [
            equipamento.linha.codigo,
            equipamento.codigo,
            sensor.codigo if sensor else '',
            sensor.nome if sensor else '',
            sensor.tipo if sensor else '',
            sensor.tag_influxdb if sensor else '',
            sensor.unidade if sensor else '',
            sensor.ativo if sensor else '',
            sensor.golden_state if sensor else '',
            sensor.valor_min if sensor else '',
            sensor.valor_max if sensor else '',
            sensor.lsl if sensor else '',
            sensor.usl if sensor else '',
            sensor.nominal if sensor else '',
            sensor.observacoes if sensor else '',
        ]

    def _worksheet_rows(self, worksheet, expected_headers):
        raw_headers = [self._text(cell.value) for cell in worksheet[1]]
        if len(raw_headers) != len(set(raw_headers)):
            raise ValidationError('A planilha possui cabecalhos duplicados.')

        missing = [header for header in expected_headers if header not in raw_headers]
        if missing:
            raise ValidationError(f'Colunas obrigatorias ausentes: {", ".join(missing)}.')

        rows = []
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            row = {
                header: values[index] if index < len(values) else None
                for index, header in enumerate(raw_headers)
            }
            rows.append((row_number, row))
        return rows

    def _validate_identity(self, row, equipamento):
        linha_codigo = self._text(row.get('linha_codigo'))
        equipamento_codigo = self._text(row.get('equipamento_codigo'))
        if linha_codigo != equipamento.linha.codigo:
            raise ValidationError(
                f'linha_codigo "{linha_codigo}" nao corresponde a {equipamento.linha.codigo}.'
            )
        if equipamento_codigo != equipamento.codigo:
            raise ValidationError(
                f'equipamento_codigo "{equipamento_codigo}" nao corresponde a {equipamento.codigo}.'
            )

    def _upsert_tag(self, row, equipamento, standard):
        nome_metrica = self._text(row.get('nome_metrica'))
        if not nome_metrica:
            raise ValidationError('nome_metrica e obrigatorio.')

        is_standard = nome_metrica in DEFAULT_TAGS_BY_NAME
        if standard and not is_standard:
            raise ValidationError(
                f'"{nome_metrica}" nao e uma variavel padrao do sistema.'
            )
        if not standard and is_standard:
            raise ValidationError(
                f'"{nome_metrica}" pertence a aba Variaveis padrao.'
            )

        tag = TagColeta.objects.filter(
            equipamento=equipamento,
            nome_metrica=nome_metrica,
        ).first()
        created = tag is None
        if created:
            tag = TagColeta(equipamento=equipamento, nome_metrica=nome_metrica)

        tag.node_id = self._text(row.get('node_id'))
        if not standard:
            tag.tipo_dado = self._text(row.get('tipo_dado')) or tag.tipo_dado
            tag.unidade = self._text(row.get('unidade'))
            tag.fator_conversao = self._to_float(
                row.get('fator_conversao'), tag.fator_conversao,
            )
        tag.formato = self._to_decimal(row.get('formato'), None)
        tag.ativa = self._to_bool(row.get('ativa'), tag.ativa)
        tag.golden_state = self._to_bool(row.get('golden_state'), tag.golden_state)
        tag.full_clean()
        tag.save()
        return 'created' if created else 'updated'

    def _upsert_sensor(self, row, equipamento):
        codigo = self._text(row.get('codigo'))
        sensor = equipamento.sensores.filter(codigo=codigo).first() if codigo else None
        created = sensor is None
        if created:
            sensor = Sensor(equipamento=equipamento, codigo=codigo)

        sensor.linha = None
        sensor.nome = self._text(row.get('nome'))
        sensor.tipo = self._text(row.get('tipo'))
        sensor.tag_influxdb = self._text(row.get('tag_influxdb'))
        if not sensor.nome or not sensor.tipo or not sensor.tag_influxdb:
            raise ValidationError('nome, tipo e tag_influxdb sao obrigatorios para sensores.')

        sensor.unidade = self._text(row.get('unidade'))
        sensor.ativo = self._to_bool(row.get('ativo'), sensor.ativo)
        sensor.golden_state = self._to_bool(row.get('golden_state'), sensor.golden_state)
        sensor.valor_min = self._to_float(row.get('valor_min'), None)
        sensor.valor_max = self._to_float(row.get('valor_max'), None)
        sensor.lsl = self._to_float(row.get('lsl'), None)
        sensor.usl = self._to_float(row.get('usl'), None)
        sensor.nominal = self._to_float(row.get('nominal'), None)
        sensor.observacoes = self._text(row.get('observacoes'))
        sensor.full_clean()
        sensor.save()
        return 'created' if created else 'updated'

    def _text(self, value):
        return '' if value is None else str(value).strip()

    def _to_float(self, value, default):
        if value in (None, ''):
            return default
        return float(str(value).replace(',', '.'))

    def _to_decimal(self, value, default):
        if value in (None, ''):
            return default
        try:
            return Decimal(str(value).replace(',', '.'))
        except InvalidOperation as exc:
            raise ValidationError(f'Valor decimal invalido: {value}') from exc

    def _to_bool(self, value, default):
        if value in (None, ''):
            return default
        normalized = str(value).strip().lower()
        if normalized in self.TRUE_VALUES:
            return True
        if normalized in self.FALSE_VALUES:
            return False
        raise ValidationError(f'Valor booleano invalido: {value}')

class SensorResource(resources.ModelResource):
    linha = fields.Field(
        column_name='linha',
        attribute='linha',
        widget=LinhaProducaoWidget(),
    )
    equipamento = fields.Field(
        column_name='equipamento',
        attribute='equipamento',
        widget=EquipamentoHierarquicoWidget(),
    )

    class Meta:
        model = Sensor
        import_id_fields = ('linha', 'equipamento', 'codigo')
        fields = (
            'codigo', 'nome', 'tipo', 'tag_influxdb', 'unidade',
            'linha', 'equipamento', 'valor_min', 'valor_max', 'ativo'
        )

class ConexaoOPCResource(resources.ModelResource):
    class Meta:
        model = ConexaoOPC
        import_id_fields = ('nome',)
        fields = (
            'nome', 'url_servidor', 'namespace_prefix', 'ativa', 'timeout'
        )

class OrdemProducaoResource(resources.ModelResource):
    linha = fields.Field(
        column_name='linha',
        attribute='linha',
        widget=LinhaProducaoWidget(),
    )
    produto = fields.Field(
        column_name='produto',
        attribute='produto',
        widget=ProdutoWidget(),
    )

    class Meta:
        model = OrdemProducao
        import_id_fields = ('codigo',)
        fields = (
            'codigo', 'linha', 'produto', 'status', 'meta_total',
            'formato_gramas', 'cuc', 'eficiencia_planejada',
            'data_planejada_inicio', 'data_inicio_real', 'data_fim_real',
            'descricao'
        )

class TurnoProducaoResource(resources.ModelResource):
    class Meta:
        model = TurnoProducao
        import_id_fields = ('codigo',)
        fields = (
            'codigo', 'nome', 'hora_inicio', 'hora_fim', 'duracao_horas', 'ativo'
        )
