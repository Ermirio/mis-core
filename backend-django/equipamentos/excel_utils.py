"""
Módulo utilitário para importar e exportar dados em Excel.

Fornece funções genéricas para:
- Exportar dados de qualquer modelo Django para Excel
- Importar dados de Excel para qualquer modelo Django
- Validação de dados durante importação
- Tratamento de erros e feedback ao usuário
"""

import io
import logging
from datetime import datetime
from typing import List, Dict, Any, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.db import models
from django.http import HttpResponse
from rest_framework.response import Response
from rest_framework import status

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exporta dados de modelos Django para arquivos Excel."""

    @staticmethod
    def export_model_to_excel(
        queryset,
        model_class,
        filename: str = None,
        fields: List[str] = None,
        exclude_fields: List[str] = None
    ) -> HttpResponse:
        """
        Exporta um queryset para Excel.

        Args:
            queryset: QuerySet do Django
            model_class: Classe do modelo
            filename: Nome do arquivo (padrão: nome_modelo_data.xlsx)
            fields: Lista de campos a exportar (padrão: todos)
            exclude_fields: Lista de campos a excluir

        Returns:
            HttpResponse com arquivo Excel
        """
        try:
            # Determina o nome do arquivo
            if not filename:
                model_name = model_class.__name__.lower()
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{model_name}_{timestamp}.xlsx"

            # Cria workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = model_class.__name__[:31]  # Excel limita a 31 caracteres

            # Obtém campos do modelo
            if fields is None:
                fields = [f.name for f in model_class._meta.get_fields()]

            if exclude_fields:
                fields = [f for f in fields if f not in exclude_fields]

            # Escreve cabeçalho
            for col_idx, field_name in enumerate(fields, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = field_name
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Escreve dados
            for row_idx, obj in enumerate(queryset, 2):
                for col_idx, field_name in enumerate(fields, 1):
                    try:
                        value = getattr(obj, field_name)
                        # Trata relacionamentos
                        if isinstance(value, models.Model):
                            value = str(value)
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    except AttributeError:
                        pass

            # Ajusta largura das colunas
            for col_idx in range(1, len(fields) + 1):
                column_letter = get_column_letter(col_idx)
                ws.column_dimensions[column_letter].width = 20

            # Cria resposta HTTP
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            wb.save(response)
            return response

        except Exception as e:
            logger.error(f"Erro ao exportar {model_class.__name__}: {e}")
            raise


class ExcelImporter:
    """Importa dados de arquivos Excel para modelos Django."""

    @staticmethod
    def import_excel_to_model(
        file,
        model_class,
        field_mapping: Dict[str, str] = None,
        skip_errors: bool = False
    ) -> Tuple[int, List[str]]:
        """
        Importa dados de Excel para um modelo Django.

        Args:
            file: Arquivo Excel (BytesIO ou arquivo)
            model_class: Classe do modelo
            field_mapping: Mapeamento de colunas Excel para campos do modelo
            skip_errors: Se True, continua mesmo com erros

        Returns:
            Tupla (quantidade_importada, lista_de_erros)
        """
        errors = []
        imported_count = 0

        try:
            # Carrega workbook
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            # Obtém cabeçalho
            header = []
            for cell in ws[1]:
                if cell.value:
                    header.append(cell.value)

            # Cria mapeamento padrão se não fornecido
            if field_mapping is None:
                field_mapping = {col: col for col in header}

            # Importa linhas
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
                try:
                    data = {}
                    for col_idx, cell in enumerate(row):
                        if col_idx < len(header):
                            excel_field = header[col_idx]
                            model_field = field_mapping.get(excel_field, excel_field)
                            if cell.value is not None:
                                data[model_field] = cell.value

                    # Cria objeto
                    if data:
                        obj = model_class(**data)
                        obj.full_clean()  # Valida
                        obj.save()
                        imported_count += 1

                except Exception as e:
                    error_msg = f"Linha {row_idx}: {str(e)}"
                    errors.append(error_msg)
                    logger.warning(error_msg)
                    if not skip_errors:
                        raise

            return imported_count, errors

        except Exception as e:
            logger.error(f"Erro ao importar para {model_class.__name__}: {e}")
            raise


def export_model_view(request, model_class, queryset=None):
    """View genérica para exportar modelo para Excel."""
    try:
        if queryset is None:
            queryset = model_class.objects.all()

        return ExcelExporter.export_model_to_excel(queryset, model_class)

    except Exception as e:
        return Response(
            {"error": f"Erro ao exportar: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


def import_model_view(request, model_class):
    """View genérica para importar modelo de Excel."""
    try:
        if "file" not in request.FILES:
            return Response(
                {"error": "Nenhum arquivo fornecido"},
                status=status.HTTP_400_BAD_REQUEST
            )

        file = request.FILES["file"]
        imported, errors = ExcelImporter.import_excel_to_model(file, model_class)

        return Response({
            "status": "success",
            "imported": imported,
            "errors": errors
        })

    except Exception as e:
        logger.error(f"Erro ao importar para {model_class.__name__}: {e}")
        return Response(
            {"error": f"Erro ao importar: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
